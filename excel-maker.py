import os
import re
import openpyxl
import docx
from pdf2docx import Converter

pdf_files = [
    os.path.join("Information", "Saved Results", "1.pdf"),
    os.path.join("Information", "Saved Results", "2.pdf"),
    os.path.join("Information", "Saved Results", "3.pdf")
]

def get_next_distinct(cells_matrix, keyword):
    for row in cells_matrix:
        if keyword in row:
            idx = row.index(keyword)
            # Find the first cell after keyword that is not the keyword
            for i in range(idx, len(row)):
                if row[i] != keyword and row[i].strip():
                    return row[i].strip()
    return ""

def extract_from_docx(docx_file):
    doc = docx.Document(docx_file)
    student_data_list = []
    all_subject_codes = set()
    
    for table in doc.tables:
        cells_matrix = []
        for row in table.rows:
            # Replaces newlines with spaces and truncates whitespaces
            row_cells = [c.text.strip().replace('\n', ' ') for c in row.cells]
            cells_matrix.append(row_cells)
            
        name = get_next_distinct(cells_matrix, "Name")
        roll = get_next_distinct(cells_matrix, "Roll No")
        
        if not roll:
            continue
            
        sgpa = ""
        cgpa = ""
        result = "FAILED"
        carry = "-"
        
        for row in cells_matrix:
            # Check unique cells in the row
            for cell in set(row):
                if "Result :" in cell:
                    result = cell.split("Result :")[1].strip()
                if "(SGPA) :" in cell:
                    try:
                        sgpa = float(cell.split("(SGPA) :")[1].strip())
                    except:
                        pass
                if "(CGPA) :" in cell:
                    try:
                        cgpa = float(cell.split("(CGPA) :")[1].strip())
                    except:
                        pass
                if "Carry Over Paper :" in cell:
                    val = cell.split("Carry Over Paper :")[1].strip().rstrip(",")
                    if val:
                        carry = val
                        
        subjects = {}
        obt_idx = None
        for row_cells in cells_matrix:
            if "Obt. Marks" in row_cells:
                obt_idx = row_cells.index("Obt. Marks")
                
            if obt_idx is not None:
                code = row_cells[0]
                # If there's a valid subject code
                if re.match(r"^[A-Z]{1,6}\d+[A-Za-z0-9\-\*]*$", code):
                    if obt_idx < len(row_cells):
                        marks = row_cells[obt_idx]
                        if marks:
                            subjects[code] = marks
                            
        all_subject_codes.update(subjects.keys())
        
        student_data_list.append({
            "roll": roll,
            "name": name,
            "sgpa": sgpa,
            "cgpa": cgpa,
            "result": result,
            "carry": carry,
            "subjects": subjects
        })
        
    return student_data_list, all_subject_codes

def parse_mark(mark_str):
    try:
        return float(mark_str)
    except:
        return -1

def main():
    os.makedirs(os.path.join("Information", "pdf2docx"), exist_ok=True)
    os.makedirs(os.path.join("Information", "docx2xlsx"), exist_ok=True)
    
    wb_all = openpyxl.Workbook()
    if wb_all.active:
        wb_all.remove(wb_all.active)
        
    for pdf_file in pdf_files:
        if not os.path.exists(pdf_file):
            print(f"File not found: {pdf_file}")
            continue
            
        docx_file = os.path.join("Information", "pdf2docx", os.path.basename(pdf_file).replace('.pdf', '.docx'))
        
        # Convert to DOCX if not present
        if not os.path.exists(docx_file):
            print(f"Converting {pdf_file} to {docx_file} using pdf2docx...")
            try:
                cv = Converter(pdf_file)
                cv.convert(docx_file, start=0, end=None)
                cv.close()
            except Exception as e:
                print(f"Failed to convert {pdf_file} to DOCX. Error: {e}")
                continue
            
        print(f"Processing {docx_file}...")
        try:
            student_data_list, all_subject_codes = extract_from_docx(docx_file)
        except Exception as e:
            print(f"Error reading {docx_file}: {e}")
            continue
        
        if not student_data_list:
            print(f"No student data found in {docx_file}.")
            continue
            
        # Overall ranks per semester
        student_data_list.sort(key=lambda x: (x["sgpa"] if isinstance(x["sgpa"], (int, float)) else 0), reverse=True)
        sorted_subject_codes = sorted(list(all_subject_codes))
        
        ws_all = wb_all.create_sheet(title=os.path.basename(docx_file).replace('.docx', ''))
        
        is_sem_1 = "1" in os.path.basename(docx_file).lower()
        
        if is_sem_1:
            headers = ["Rank", "Roll Number", "Student's Name", "SGPA", "Result"] + sorted_subject_codes + ["Carry Over Paper"]
        else:
            headers = ["Rank", "Roll Number", "Student's Name", "SGPA", "CGPA", "Result"] + sorted_subject_codes + ["Carry Over Paper"]
            
        ws_all.append(headers)
        
        for rank, student in enumerate(student_data_list, 1):
            if is_sem_1:
                row_data = [
                    rank,
                    student["roll"],
                    student["name"],
                    student["sgpa"],
                    student["result"]
                ]
            else:
                row_data = [
                    rank,
                    student["roll"],
                    student["name"],
                    student["sgpa"],
                    student["cgpa"],
                    student["result"]
                ]
                
            for code in sorted_subject_codes:
                mark = student["subjects"].get(code, "")
                row_data.append(mark)
                
            row_data.append(student["carry"])
            ws_all.append(row_data)
            
        # Group students by subject
        subject_students_map = {}
        for student in student_data_list:
            for code, marks in student["subjects"].items():
                if code not in subject_students_map:
                    subject_students_map[code] = []
                subject_students_map[code].append({
                    "roll": student["roll"],
                    "name": student["name"],
                    "marks": marks
                })
        
        wb_subjects = openpyxl.Workbook()
        if wb_subjects.active:
            wb_subjects.remove(wb_subjects.active)
            
        sorted_subjects = sorted(list(subject_students_map.keys()))
        
        for code in sorted_subjects:
            safe_sheet_name = re.sub(r'[\\*?:/\[\]]', '_', code)[:31]
            ws_subj = wb_subjects.create_sheet(title=safe_sheet_name)
            
            headers = ["Subject Rank", "Roll Number", "Student's Name", "Marks Obtained"]
            ws_subj.append(headers)
            
            students_in_subject = subject_students_map[code]
            students_in_subject.sort(key=lambda x: parse_mark(x["marks"]), reverse=True)
            
            for rank, student in enumerate(students_in_subject, 1):
                ws_subj.append([
                    rank,
                    student["roll"],
                    student["name"],
                    student["marks"]
                ])
                
        if wb_subjects.sheetnames:
            base_name = os.path.basename(docx_file).replace('.docx', '')
            output_file_subjects = os.path.join("Information", "docx2xlsx", f"{base_name}_subject_ranks.xlsx")
            wb_subjects.save(output_file_subjects)
            print(f"✅ Saved subject-wise ranking for {base_name} as {output_file_subjects}")
        else:
            print(f"No subject data found for {docx_file}, skipping subject workbook.")

    if wb_all.sheetnames:
        output_file_all = os.path.join("Information", "docx2xlsx", "all_results.xlsx")
        wb_all.save(output_file_all)
        print(f"\n✅ Consolidated Excel file saved as {output_file_all}\n")
    else:
        print("\n❌ No data was found across any files. No master Excel generated.\n")

if __name__ == "__main__":
    main()
