import os
import re
import docx
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment

docx_files = [
    os.path.join("Information", "pdf2docx", "1.docx"),
    os.path.join("Information", "pdf2docx", "2.docx"),
    os.path.join("Information", "pdf2docx", "3.docx")
]

def get_next_distinct(cells_matrix, keyword):
    for row in cells_matrix:
        if keyword in row:
            try:
                idx = row.index(keyword)
                # Find the first cell after keyword that is not the keyword
                for i in range(idx + 1, len(row)):
                    if row[i] != keyword and row[i].strip():
                        return row[i].strip()
            except ValueError:
                continue
    return ""

def extract_from_docx(docx_file):
    doc = docx.Document(docx_file)
    student_data_list = []
    
    for table in doc.tables:
        cells_matrix = []
        for row in table.rows:
            row_cells = [c.text.strip().replace('\n', ' ') for c in row.cells]
            cells_matrix.append(row_cells)
            
        name = get_next_distinct(cells_matrix, "Name")
        roll = get_next_distinct(cells_matrix, "Roll No")
        
        if not roll:
            continue
            
        subjects = {}
        obt_idx = None
        for row_cells in cells_matrix:
            if "Obt. Marks" in row_cells:
                try:
                    obt_idx = row_cells.index("Obt. Marks")
                except ValueError:
                    pass
                
            if obt_idx is not None:
                code = row_cells[0]
                if re.match(r"^[A-Z]{1,6}\d+[A-Za-z0-9\-\*]*$", code):
                    if obt_idx < len(row_cells):
                        marks = row_cells[obt_idx]
                        if marks:
                            subjects[code] = marks
                            
        student_data_list.append({
            "roll": roll,
            "name": name,
            "subjects": subjects
        })
        
    return student_data_list

def parse_mark(mark_str):
    try:
        clean_str = re.sub(r'[^\d.]', '', str(mark_str))
        if not clean_str:
            return 0.0
        return float(clean_str)
    except Exception:
        return 0.0

def main():
    # 1. Load subjects and credits
    subjects_dir = os.path.join("Information", "Subjects")
    if not os.path.exists(subjects_dir):
        print(f"Subjects directory not found: {subjects_dir}")
        return
        
    subject_files = [f for f in os.listdir(subjects_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    # Structure: SubjectName -> { "Course_Code": credits }
    subject_credits = {}
    
    for file in subject_files:
        subj_name = file.replace('.xlsx', '')
        filepath = os.path.join(subjects_dir, file)
        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            print(f"Error reading {filepath}: {e}")
            continue
            
        credits_map = {}
        if 'Course Code' in df.columns and 'Credits' in df.columns:
            for _, row in df.iterrows():
                code = str(row['Course Code']).strip()
                try:
                    # Clean credits if needed
                    cred_str = re.sub(r'[^\d.]', '', str(row['Credits']))
                    cred = float(cred_str)
                    credits_map[code] = cred
                except Exception:
                    pass
            
            if credits_map:
                subject_credits[subj_name] = credits_map

    # 2. Extract all students data per semester
    # Structure: roll -> { "name": ..., "sems": { 1: {code: marks}, 2: {code: marks}, ... } }
    students_all = {}
    sem_keys = []
    
    for i, docx_file in enumerate(docx_files):
        sem = i + 1
        if not os.path.exists(docx_file):
            print(f"File not found: {docx_file}")
            continue
            
        print(f"Processing semester {sem} file: {docx_file}...")
        sem_keys.append(sem)
        sem_data = extract_from_docx(docx_file)
        
        for student in sem_data:
            roll = student["roll"]
            name = student["name"]
            
            if roll not in students_all:
                students_all[roll] = {"name": name, "sems": {}}
                
            students_all[roll]["sems"][sem] = student["subjects"]

    if not students_all:
        print("No student data found.")
        return

    # 3. Process each subject
    out_dir = os.path.join("Information", "Subject_Rankings")
    os.makedirs(out_dir, exist_ok=True)
    
    arial_14 = Font(name='Arial', size=14)
    arial_14_bold = Font(name='Arial', size=14, bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center')

    for subj_name, credits_map in subject_credits.items():
        print(f"Generating rankings for subject: {subj_name}...")
        
        # PRE-PASS: Determine which codes are present in each semester for this subject
        codes_in_sem = {sem: set() for sem in sem_keys}
        all_codes_for_subj = set()
        for roll, data in students_all.items():
            for sem in sem_keys:
                if sem in data["sems"]:
                    for code in data["sems"][sem]:
                        if code in credits_map:
                            codes_in_sem[sem].add(code)
                            all_codes_for_subj.add(code)

        # We will collect records of students who took at least one paper in this subject.
        overall_records = []
        sem_records = {sem: [] for sem in sem_keys}

        for roll, data in students_all.items():
            name = data["name"]
            
            total_sum_product = 0.0
            total_sum_credits = 0.0
            overall_papers = {}
            
            took_subject_overall = False
            absent_any_sem = False
            
            for sem in sem_keys:
                sem_sum_product = 0.0
                sem_sum_credits = 0.0
                sem_papers = {}
                took_in_sem = False
                
                subject_offered_in_sem = len(codes_in_sem[sem]) > 0
                
                if sem in data["sems"]:
                    for code, mark_str in data["sems"][sem].items():
                        if code in credits_map:
                            took_subject_overall = True
                            took_in_sem = True
                            
                            sem_papers[code] = mark_str
                            overall_papers[code] = mark_str
                            
                            mark_val = parse_mark(mark_str)
                            cred_val = credits_map[code]
                            
                            sem_sum_product += mark_val * cred_val
                            sem_sum_credits += cred_val
                            
                            if str(mark_str).strip().upper() in ['AB', 'ABS', 'A', '---', '', 'F']:
                                absent_any_sem = True
                                
                if subject_offered_in_sem and not took_in_sem:
                    absent_any_sem = True
                elif subject_offered_in_sem and sem not in data["sems"]:
                    absent_any_sem = True
                    
                if took_in_sem and sem_sum_credits > 0:
                    sem_pct = sem_sum_product / sem_sum_credits
                    
                    total_sum_product += sem_sum_product
                    total_sum_credits += sem_sum_credits
                    
                    sem_records[sem].append({
                        "roll": roll,
                        "name": name,
                        "pct": sem_pct,
                        "papers": sem_papers
                    })
            
            if took_subject_overall:
                overall_pct = 0.0
                if total_sum_credits > 0:
                    overall_pct = total_sum_product / total_sum_credits
                overall_records.append({
                    "roll": roll,
                    "name": name,
                    "pct": overall_pct,
                    "papers": overall_papers,
                    "absent_any_sem": absent_any_sem
                })

                
        # If nobody took this subject, skip
        if not overall_records:
            print(f"  No student data for {subj_name}. Skipping.\n")
            continue
            
        # Create Excel file
        wb = openpyxl.Workbook()
        if wb.active:
            wb.remove(wb.active)
            
        def apply_styles_and_width(ws):
            for row in ws.iter_rows():
                for cell in row:
                    if cell.row == 1:
                        cell.font = arial_14_bold
                    else:
                        cell.font = arial_14
                    cell.alignment = center_aligned
                    
            for col in ws.columns:
                max_length = 0
                if not col: continue
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        v = str(cell.value) if cell.value is not None else ""
                        if len(v) > max_length:
                            max_length = len(v)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(60, (max_length + 2) * 1.5)

        # 1. Overall Toppers Sheet
        overall_records.sort(key=lambda x: (
            0 if x["absent_any_sem"] else 1,
            x["pct"]
        ), reverse=True)
        
        sorted_all_codes = sorted(list(all_codes_for_subj))
        
        # Avoid invalid characters in sheet name and limit to 31 chars
        ws_overall = wb.create_sheet(title="Overall Toppers")
        headers_overall = ["Rank", "Roll Number", "Student's Name", "Overall Percentage"] + sorted_all_codes
        ws_overall.append(headers_overall)
        
        current_rank = 1
        for rec in overall_records:
            if rec["absent_any_sem"]:
                rank_str = "N.A."
            else:
                rank_str = current_rank
                current_rank += 1
                
            row_data = [
                rank_str,
                rec["roll"],
                rec["name"],
                round(rec["pct"], 2)
            ]
            for code in sorted_all_codes:
                row_data.append(rec["papers"].get(code, ""))
            ws_overall.append(row_data)
            
        apply_styles_and_width(ws_overall)
        
        # 2. Semester Sheets
        for sem in sem_keys:
            if not sem_records[sem]:
                continue
                
            sorted_sem_codes = sorted(list(codes_in_sem[sem]))
            ws_sem = wb.create_sheet(title=f"Semester {sem} Toppers")
            
            headers_sem = ["Rank", "Roll Number", "Student's Name", f"Semester {sem} Percentage"] + sorted_sem_codes
            ws_sem.append(headers_sem)
            
            sem_records[sem].sort(key=lambda x: x["pct"], reverse=True)
            
            for rank, rec in enumerate(sem_records[sem], 1):
                row_data = [
                    rank,
                    rec["roll"],
                    rec["name"],
                    round(rec["pct"], 2)
                ]
                for code in sorted_sem_codes:
                    row_data.append(rec["papers"].get(code, ""))
                ws_sem.append(row_data)
                
            apply_styles_and_width(ws_sem)

        output_file = os.path.join(out_dir, f"{subj_name}_Rankings.xlsx")
        wb.save(output_file)
        print(f"[SUCCESS] Saved {subj_name} rankings to {output_file}\n")

if __name__ == "__main__":
    main()
