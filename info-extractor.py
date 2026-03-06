import docx
import re
import openpyxl
import os

docx_files = [
    os.path.join("Information", "pdf2docx", "1.docx"),
    os.path.join("Information", "pdf2docx", "2.docx"),
    os.path.join("Information", "pdf2docx", "3.docx")
]

def get_next_distinct(cells_matrix, keyword):
    for row in cells_matrix:
        if keyword in row:
            idx = row.index(keyword)
            # Find the first cell after keyword that is not the keyword
            for i in range(idx, len(row)):
                if row[i] != keyword and row[i].strip():
                    return row[i].strip()
    return ""

def extract_from_docx(docx_file):
    doc = docx.Document(docx_file)
    student_data_list = []
    all_subject_codes = set()
    
    for table in doc.tables:
        cells_matrix = []
        for row in table.rows:
            # Replaces newlines with spaces and truncates whitespaces
            row_cells = [c.text.strip().replace('\n', ' ') for c in row.cells]
            cells_matrix.append(row_cells)
            
        name = get_next_distinct(cells_matrix, "Name")
        roll = get_next_distinct(cells_matrix, "Roll No")
        
        if not roll:
            continue
            
        sgpa = ""
        cgpa = ""
        result = "FAILED"
        carry = "-"
        
        for row in cells_matrix:
            # Check unique cells in the row
            for cell in set(row):
                if "Result :" in cell:
                    result = cell.split("Result :")[1].strip()
                if "(SGPA) :" in cell:
                    try:
                        sgpa = float(cell.split("(SGPA) :")[1].strip())
                    except:
                        pass
                if "(CGPA) :" in cell:
                    try:
                        cgpa = float(cell.split("(CGPA) :")[1].strip())
                    except:
                        pass
                if "Carry Over Paper :" in cell:
                    val = cell.split("Carry Over Paper :")[1].strip().rstrip(",")
                    if val:
                        carry = val
                        
        subjects = {}
        obt_idx = None
        for row_cells in cells_matrix:
            if "Obt. Marks" in row_cells:
                obt_idx = row_cells.index("Obt. Marks")
                
            if obt_idx is not None:
                code = row_cells[0]
                # If there's a valid subject code
                if re.match(r"^[A-Z]{1,6}\d+[A-Za-z0-9\-\*]*$", code):
                    if obt_idx < len(row_cells):
                        marks = row_cells[obt_idx]
                        if marks:
                            subjects[code] = marks
                            
        all_subject_codes.update(subjects.keys())
        
        student_data_list.append({
            "roll": roll,
            "name": name,
            "sgpa": sgpa,
            "cgpa": cgpa,
            "result": result,
            "carry": carry,
            "subjects": subjects
        })
        
    return student_data_list, all_subject_codes

def main():
    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)
        
    os.makedirs(os.path.join("Information", "docx2xlsx"), exist_ok=True)
    
    for docx_file in docx_files:
        if not os.path.exists(docx_file):
            print(f"File not found: {docx_file}")
            continue
            
        print(f"Processing {docx_file}...")
        student_data_list, all_subject_codes = extract_from_docx(docx_file)
        
        if not student_data_list:
            print(f"No student data found in {docx_file}.")
            continue
            
        student_data_list.sort(key=lambda x: (x["sgpa"] if isinstance(x["sgpa"], (int, float)) else 0), reverse=True)
        sorted_subject_codes = sorted(list(all_subject_codes))
        
        ws = wb.create_sheet(title=os.path.basename(docx_file).replace('.docx', ''))
        
        is_sem_1 = "1.docx" in docx_file.lower()
        
        if is_sem_1:
            headers = ["Rank", "Roll Number", "Student's Name", "SGPA", "Result"] + sorted_subject_codes + ["Carry Over Paper"]
        else:
            headers = ["Rank", "Roll Number", "Student's Name", "SGPA", "CGPA", "Result"] + sorted_subject_codes + ["Carry Over Paper"]
            
        ws.append(headers)
        
        for rank, student in enumerate(student_data_list, 1):
            if is_sem_1:
                row_data = [
                    rank,
                    student["roll"],
                    student["name"],
                    student["sgpa"],
                    student["result"]
                ]
            else:
                row_data = [
                    rank,
                    student["roll"],
                    student["name"],
                    student["sgpa"],
                    student["cgpa"],
                    student["result"]
                ]
                
            for code in sorted_subject_codes:
                mark = student["subjects"].get(code, "")
                row_data.append(mark)
                
            row_data.append(student["carry"])
            ws.append(row_data)
            
    if not wb.sheetnames:
        print("No student data to write; no Excel file created.")
        return
        
    output_file = os.path.join("Information", "docx2xlsx", "all_results_from_docx.xlsx")
    wb.save(output_file)
    print(f"✅ Consolidated Excel file saved as {output_file}\n")

if __name__ == "__main__":
    main()
