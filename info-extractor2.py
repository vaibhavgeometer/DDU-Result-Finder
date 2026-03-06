import docx
import re
import openpyxl
import os

docx_files = [
    os.path.join("Information", "pdf2docx", "1.docx"),
    os.path.join("Information", "pdf2docx", "2.docx"),
    os.path.join("Information", "pdf2docx", "3.docx")
]

def get_next_distinct(cells_matrix, keyword):
    for row in cells_matrix:
        if keyword in row:
            idx = row.index(keyword)
            # Find the first cell after keyword that is not the keyword
            for i in range(idx, len(row)):
                if row[i] != keyword and row[i].strip():
                    return row[i].strip()
    return ""

def extract_from_docx(docx_file):
    doc = docx.Document(docx_file)
    student_data_list = []
    
    for table in doc.tables:
        cells_matrix = []
        for row in table.rows:
            # Replaces newlines with spaces and truncates whitespaces
            row_cells = [c.text.strip().replace('\n', ' ') for c in row.cells]
            cells_matrix.append(row_cells)
            
        name = get_next_distinct(cells_matrix, "Name")
        roll = get_next_distinct(cells_matrix, "Roll No")
        
        if not roll:
            continue
                        
        subjects = {}
        obt_idx = None
        for row_cells in cells_matrix:
            if "Obt. Marks" in row_cells:
                obt_idx = row_cells.index("Obt. Marks")
                
            if obt_idx is not None:
                code = row_cells[0]
                # If there's a valid subject code
                if re.match(r"^[A-Z]{1,6}\d+[A-Za-z0-9\-\*]*$", code):
                    if obt_idx < len(row_cells):
                        marks = row_cells[obt_idx]
                        if marks:
                            subjects[code] = marks
                            
        student_data_list.append({
            "roll": roll,
            "name": name,
            "subjects": subjects
        })
        
    return student_data_list

def parse_mark(mark_str):
    try:
        return float(mark_str)
    except:
        return -1

def main():
    os.makedirs(os.path.join("Information", "docx2xlsx"), exist_ok=True)
    
    for docx_file in docx_files:
        if not os.path.exists(docx_file):
            print(f"File not found: {docx_file}")
            continue
            
        print(f"Processing {docx_file}...")
        student_data_list = extract_from_docx(docx_file)
        
        if not student_data_list:
            print(f"No student data found in {docx_file}.")
            continue
            
        # Group students by subject
        subject_students_map = {}
        for student in student_data_list:
            for code, marks in student["subjects"].items():
                if code not in subject_students_map:
                    subject_students_map[code] = []
                subject_students_map[code].append({
                    "roll": student["roll"],
                    "name": student["name"],
                    "marks": marks
                })
        
        # Create an Excel file for this semester
        wb = openpyxl.Workbook()
        if wb.active:
            wb.remove(wb.active)
            
        sorted_subjects = sorted(list(subject_students_map.keys()))
        
        for code in sorted_subjects:
            # Sanitize sheet name if it exceeds 31 chars or has invalid chars
            safe_sheet_name = re.sub(r'[\\*?:/\[\]]', '_', code)[:31]
            ws = wb.create_sheet(title=safe_sheet_name)
            
            headers = ["Subject Rank", "Roll Number", "Student's Name", "Marks Obtained"]
            ws.append(headers)
            
            # Sort students by marks descending
            students_in_subject = subject_students_map[code]
            students_in_subject.sort(key=lambda x: parse_mark(x["marks"]), reverse=True)
            
            for rank, student in enumerate(students_in_subject, 1):
                ws.append([
                    rank,
                    student["roll"],
                    student["name"],
                    student["marks"]
                ])
                
        if not wb.sheetnames:
            print(f"No valid subjects extracted for {docx_file}, skipping Excel creation.")
            continue
            
        base_name = os.path.basename(docx_file).replace('.docx', '')
        output_file = os.path.join("Information", "docx2xlsx", f"{base_name}_subject_ranks.xlsx")
        wb.save(output_file)
        print(f"✅ Saved subject-wise ranking for {base_name} as {output_file}\n")

if __name__ == "__main__":
    main()
